import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

# Define industries and their specific metrics with ranges
industry_data = {
    "General": [
        {"Metric": "P/E Ratio", "Ideal Range": "10-20", "Significantly Undervalued": "<10", "Slightly Undervalued": "10-12", "Fairly Valued": "12-18", "Slightly Overvalued": "18-25", "Significantly Overvalued": ">25"},
        {"Metric": "PEG Ratio", "Ideal Range": "0.5-1.0", "Significantly Undervalued": "<0.5", "Slightly Undervalued": "0.5-0.75", "Fairly Valued": "0.75-1.0", "Slightly Overvalued": "1.0-2.0", "Significantly Overvalued": ">2.0"},
        {"Metric": "P/B Ratio", "Ideal Range": "1-3", "Significantly Undervalued": "<1", "Slightly Undervalued": "1-1.5", "Fairly Valued": "1.5-2.5", "Slightly Overvalued": "2.5-5", "Significantly Overvalued": ">5"},
        {"Metric": "P/FCF Ratio", "Ideal Range": "10-15", "Significantly Undervalued": "<10", "Slightly Undervalued": "10-12", "Fairly Valued": "12-15", "Slightly Overvalued": "15-20", "Significantly Overvalued": ">20"},
        {"Metric": "EV/EBITDA", "Ideal Range": "5-12", "Significantly Undervalued": "<5", "Slightly Undervalued": "5-8", "Fairly Valued": "8-12", "Slightly Overvalued": "12-15", "Significantly Overvalued": ">15"},
        {"Metric": "Debt/Equity Ratio", "Ideal Range": "<0.5", "Significantly Undervalued": "N/A", "Slightly Undervalued": "<0.3", "Fairly Valued": "0.3-0.5", "Slightly Overvalued": "0.5-1.0", "Significantly Overvalued": ">1.0"},
        {"Metric": "Current Ratio", "Ideal Range": "1.5-2.5", "Significantly Undervalued": "<1.0", "Slightly Undervalued": "1.0-1.5", "Fairly Valued": "1.5-2.5", "Slightly Overvalued": "2.5-3.0", "Significantly Overvalued": ">3.0"},
        {"Metric": "ROE", "Ideal Range": "15%-25%", "Significantly Undervalued": "<10%", "Slightly Undervalued": "10%-15%", "Fairly Valued": "15%-25%", "Slightly Overvalued": "25%-30%", "Significantly Overvalued": ">30%"},
        {"Metric": "ROA", "Ideal Range": "5%-10%", "Significantly Undervalued": "<5%", "Slightly Undervalued": "5%-7.5%", "Fairly Valued": "7.5%-10%", "Slightly Overvalued": "10%-12%", "Significantly Overvalued": ">12%"},
        {"Metric": "Gross Margin", "Ideal Range": "40%-70%", "Significantly Undervalued": "<40%", "Slightly Undervalued": "40%-50%", "Fairly Valued": "50%-60%", "Slightly Overvalued": "60%-70%", "Significantly Overvalued": ">70%"},
        {"Metric": "Quick Ratio", "Ideal Range": "1.0-2.0", "Significantly Undervalued": "<0.7", "Slightly Undervalued": "0.7-1.0", "Fairly Valued": "1.0-2.0", "Slightly Overvalued": "2.0-3.0", "Significantly Overvalued": ">3.0"},
        {"Metric": "FCF Yield", "Ideal Range": "5%-10%", "Significantly Undervalued": ">10%", "Slightly Undervalued": "8%-10%", "Fairly Valued": "5%-8%", "Slightly Overvalued": "2%-5%", "Significantly Overvalued": "<2%"},
        {"Metric": "Dividend Yield", "Ideal Range": "2%-5%", "Significantly Undervalued": ">5%", "Slightly Undervalued": "4%-5%", "Fairly Valued": "2%-4%", "Slightly Overvalued": "1%-2%", "Significantly Overvalued": "<1%"},
        {"Metric": "EV/Sales", "Ideal Range": "1-3", "Significantly Undervalued": "<1", "Slightly Undervalued": "1-2", "Fairly Valued": "2-3", "Slightly Overvalued": "3-5", "Significantly Overvalued": ">5"},
        {"Metric": "Inventory Turnover", "Ideal Range": "5-10", "Significantly Undervalued": "<3", "Slightly Undervalued": "3-5", "Fairly Valued": "5-10", "Slightly Overvalued": "10-15", "Significantly Overvalued": ">15"},
        {"Metric": "Asset Turnover", "Ideal Range": "1-2", "Significantly Undervalued": "<0.5", "Slightly Undervalued": "0.5-1", "Fairly Valued": "1-2", "Slightly Overvalued": "2-3", "Significantly Overvalued": ">3"},
    ],
    "Technology": [
        {"Metric": "P/E Ratio", "Ideal Range": "15-30", "Significantly Undervalued": "<15", "Slightly Undervalued": "15-20", "Fairly Valued": "20-25", "Slightly Overvalued": "25-40", "Significantly Overvalued": ">40"},
        {"Metric": "PEG Ratio", "Ideal Range": "1-2", "Significantly Undervalued": "<1", "Slightly Undervalued": "1-1.5", "Fairly Valued": "1.5-2", "Slightly Overvalued": "2-3", "Significantly Overvalued": ">3"},
        {"Metric": "P/B Ratio", "Ideal Range": "3-10", "Significantly Undervalued": "<3", "Slightly Undervalued": "3-5", "Fairly Valued": "5-10", "Slightly Overvalued": "10-15", "Significantly Overvalued": ">15"},
        {"Metric": "P/FCF Ratio", "Ideal Range": "20-30", "Significantly Undervalued": "<20", "Slightly Undervalued": "20-25", "Fairly Valued": "25-30", "Slightly Overvalued": "30-40", "Significantly Overvalued": ">40"},
        {"Metric": "EV/EBITDA", "Ideal Range": "10-20", "Significantly Undervalued": "<10", "Slightly Undervalued": "10-15", "Fairly Valued": "15-20", "Slightly Overvalued": "20-25", "Significantly Overvalued": ">25"},
        {"Metric": "Debt/Equity Ratio", "Ideal Range": "<0.6", "Significantly Undervalued": "N/A", "Slightly Undervalued": "<0.4", "Fairly Valued": "0.4-0.6", "Slightly Overvalued": "0.6-1.2", "Significantly Overvalued": ">1.2"},
        {"Metric": "Current Ratio", "Ideal Range": "1.2-2.0", "Significantly Undervalued": "<0.8", "Slightly Undervalued": "0.8-1.2", "Fairly Valued": "1.2-2.0", "Slightly Overvalued": "2.0-2.5", "Significantly Overvalued": ">2.5"},
        {"Metric": "ROE", "Ideal Range": "20%-30%", "Significantly Undervalued": "<10%", "Slightly Undervalued": "10%-20%", "Fairly Valued": "20%-30%", "Slightly Overvalued": "30%-35%", "Significantly Overvalued": ">35%"},
        {"Metric": "ROA", "Ideal Range": "8%-15%", "Significantly Undervalued": "<5%", "Slightly Undervalued": "5%-8%", "Fairly Valued": "8%-15%", "Slightly Overvalued": "15%-20%", "Significantly Overvalued": ">20%"},
        {"Metric": "Gross Margin", "Ideal Range": "50%-80%", "Significantly Undervalued": "<40%", "Slightly Undervalued": "40%-50%", "Fairly Valued": "50%-80%", "Slightly Overvalued": "80%-85%", "Significantly Overvalued": ">85%"},
        {"Metric": "Quick Ratio", "Ideal Range": "1.0-2.5", "Significantly Undervalued": "<0.7", "Slightly Undervalued": "0.7-1.0", "Fairly Valued": "1.0-2.5", "Slightly Overvalued": "2.5-3.5", "Significantly Overvalued": ">3.5"},
        {"Metric": "FCF Yield", "Ideal Range": "10%-15%", "Significantly Undervalued": ">15%", "Slightly Undervalued": "10%-15%", "Fairly Valued": "10%-15%", "Slightly Overvalued": "5%-10%", "Significantly Overvalued": "<3%"},
        {"Metric": "Dividend Yield", "Ideal Range": "1%-3%", "Significantly Undervalued": ">4%", "Slightly Undervalued": "3%-4%", "Fairly Valued": "1%-3%", "Slightly Overvalued": "0.5%-1%", "Significantly Overvalued": "<0.5%"},
        {"Metric": "EV/Sales", "Ideal Range": "3-6", "Significantly Undervalued": "<3", "Slightly Undervalued": "3-4", "Fairly Valued": "4-6", "Slightly Overvalued": "6-10", "Significantly Overvalued": ">10"},
        {"Metric": "Inventory Turnover", "Ideal Range": "4-8", "Significantly Undervalued": "<2", "Slightly Undervalued": "2-4", "Fairly Valued": "4-8", "Slightly Overvalued": "8-12", "Significantly Overvalued": ">12"},
        {"Metric": "Asset Turnover", "Ideal Range": "1.5-3", "Significantly Undervalued": "<1", "Slightly Undervalued": "1-1.5", "Fairly Valued": "1.5-3", "Slightly Overvalued": "3-4", "Significantly Overvalued": ">4"},
    ],
    "Energy": [
        {"Metric": "P/E Ratio", "Ideal Range": "5-15", "Significantly Undervalued": "<5", "Slightly Undervalued": "5-8", "Fairly Valued": "8-15", "Slightly Overvalued": "15-20", "Significantly Overvalued": ">20"},
        {"Metric": "PEG Ratio", "Ideal Range": "0.5-1.5", "Significantly Undervalued": "<0.5", "Slightly Undervalued": "0.5-1.0", "Fairly Valued": "1.0-1.5", "Slightly Overvalued": "1.5-2.5", "Significantly Overvalued": ">2.5"},
        {"Metric": "P/B Ratio", "Ideal Range": "1-2.5", "Significantly Undervalued": "<1", "Slightly Undervalued": "1-1.5", "Fairly Valued": "1.5-2.5", "Slightly Overvalued": "2.5-4", "Significantly Overvalued": ">4"},
        {"Metric": "P/FCF Ratio", "Ideal Range": "8-18", "Significantly Undervalued": "<8", "Slightly Undervalued": "8-12", "Fairly Valued": "12-18", "Slightly Overvalued": "18-25", "Significantly Overvalued": ">25"},
        {"Metric": "EV/EBITDA", "Ideal Range": "6-14", "Significantly Undervalued": "<6", "Slightly Undervalued": "6-10", "Fairly Valued": "10-14", "Slightly Overvalued": "14-20", "Significantly Overvalued": ">20"},
        {"Metric": "Debt/Equity Ratio", "Ideal Range": "<1.0", "Significantly Undervalued": "N/A", "Slightly Undervalued": "<0.6", "Fairly Valued": "0.6-1.0", "Slightly Overvalued": "1.0-2.0", "Significantly Overvalued": ">2.0"},
        {"Metric": "Current Ratio", "Ideal Range": "1.2-2.0", "Significantly Undervalued": "<0.8", "Slightly Undervalued": "0.8-1.2", "Fairly Valued": "1.2-2.0", "Slightly Overvalued": "2.0-2.5", "Significantly Overvalued": ">2.5"},
        {"Metric": "ROE", "Ideal Range": "10%-20%", "Significantly Undervalued": "<8%", "Slightly Undervalued": "8%-10%", "Fairly Valued": "10%-20%", "Slightly Overvalued": "20%-25%", "Significantly Overvalued": ">25%"},
        {"Metric": "ROA", "Ideal Range": "4%-10%", "Significantly Undervalued": "<3%", "Slightly Undervalued": "3%-4%", "Fairly Valued": "4%-10%", "Slightly Overvalued": "10%-12%", "Significantly Overvalued": ">12%"},
        {"Metric": "Gross Margin", "Ideal Range": "30%-50%", "Significantly Undervalued": "<25%", "Slightly Undervalued": "25%-30%", "Fairly Valued": "30%-50%", "Slightly Overvalued": "50%-55%", "Significantly Overvalued": ">55%"},
        {"Metric": "Quick Ratio", "Ideal Range": "0.8-1.5", "Significantly Undervalued": "<0.5", "Slightly Undervalued": "0.5-0.8", "Fairly Valued": "0.8-1.5", "Slightly Overvalued": "1.5-2.0", "Significantly Overvalued": ">2.0"},
        {"Metric": "FCF Yield", "Ideal Range": "4%-8%", "Significantly Undervalued": ">10%", "Slightly Undervalued": "8%-10%", "Fairly Valued": "4%-8%", "Slightly Overvalued": "2%-4%", "Significantly Overvalued": "<1%"},
        {"Metric": "Dividend Yield", "Ideal Range": "3%-6%", "Significantly Undervalued": ">7%", "Slightly Undervalued": "6%-7%", "Fairly Valued": "3%-6%", "Slightly Overvalued": "2%-3%", "Significantly Overvalued": "<2%"},
        {"Metric": "EV/Sales", "Ideal Range": "1-3", "Significantly Undervalued": "<1", "Slightly Undervalued": "1-2", "Fairly Valued": "2-3", "Slightly Overvalued": "3-5", "Significantly Overvalued": ">5"},
        {"Metric": "Inventory Turnover", "Ideal Range": "3-7", "Significantly Undervalued": "<2", "Slightly Undervalued": "2-3", "Fairly Valued": "3-7", "Slightly Overvalued": "7-10", "Significantly Overvalued": ">10"},
        {"Metric": "Asset Turnover", "Ideal Range": "0.5-1.5", "Significantly Undervalued": "<0.3", "Slightly Undervalued": "0.3-0.5", "Fairly Valued": "0.5-1.5", "Slightly Overvalued": "1.5-2.0", "Significantly Overvalued": ">2.0"},
    ],
    # Add more industries as needed
}

# Define assessment categories and their colors
category_colors = {
    "Significantly Undervalued": "90EE90",      # Light Green
    "Slightly Undervalued": "98FB98",          # Pale Green
    "Fairly Valued": "FFFFFF",                 # White
    "Slightly Overvalued": "FFA07A",           # Light Salmon
    "Significantly Overvalued": "FF6347",      # Tomato
    "N/A": "D3D3D3",                           # Light Grey for Not Applicable
}

# Create a workbook and add sheets for each industry
wb = openpyxl.Workbook()

# Remove the default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Step 1: Create all worksheets and populate them with data
for industry, metrics in industry_data.items():
    ws = wb.create_sheet(title=industry)
    
    # Add headers
    headers = [
        "Metric", "Ideal Range", 
        "Significantly Undervalued", "Slightly Undervalued", 
        "Fairly Valued", "Slightly Overvalued", 
        "Significantly Overvalued", "Input Value", "Assessment"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add metric data
    for row, metric in enumerate(metrics, start=2):
        ws.cell(row=row, column=1, value=metric["Metric"])
        ws.cell(row=row, column=2, value=metric["Ideal Range"])
        ws.cell(row=row, column=3, value=metric["Significantly Undervalued"])
        ws.cell(row=row, column=4, value=metric["Slightly Undervalued"])
        ws.cell(row=row, column=5, value=metric["Fairly Valued"])
        ws.cell(row=row, column=6, value=metric["Slightly Overvalued"])
        ws.cell(row=row, column=7, value=metric["Significantly Overvalued"])
        
        # Set alignment for metric rows
        for col in range(1, 8):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
    
    # Add input and assessment sections
    for row, metric in enumerate(metrics, start=2):
        input_cell = ws.cell(row=row, column=8)
        input_cell.number_format = '0.00'
        input_cell.alignment = Alignment(horizontal="center")
        
        assessment_cell = ws.cell(row=row, column=9)
        assessment_cell.alignment = Alignment(horizontal="center")
        
        # Determine if the metric values are percentages by checking if "%" is in the ideal range
        is_percentage = "%" in metric["Ideal Range"]
        
        if is_percentage:
            # Handle percentage metrics
            if metric["Significantly Undervalued"].strip().upper() == "N/A":
                # Skip "Significantly Undervalued" and start with "Slightly Undervalued"
                try:
                    sl_und_upper = float(metric["Slightly Undervalued"].split('-')[1].replace('%','')) / 100
                    fair_upper = float(metric["Fairly Valued"].split('-')[1].replace('%','')) / 100
                    sl_over_upper = float(metric["Slightly Overvalued"].split('-')[1].replace('%','')) / 100
                    formula = f"""=IF(ISBLANK(H{row}), "",
                        IF(H{row}<{sl_und_upper}, "Slightly Undervalued",
                        IF(H{row}<={fair_upper}, "Fairly Valued",
                        IF(H{row}<{sl_over_upper}, "Slightly Overvalued",
                        "Significantly Overvalued"))))"""
                except Exception as e:
                    # If parsing fails, set a default assessment
                    formula = f"""=IF(ISBLANK(H{row}), "", "Assessment")"""
            else:
                # Regular formula with all categories
                try:
                    sig_und = float(metric["Significantly Undervalued"].replace('<','').replace('>','').replace('%','')) / 100
                    sl_und_upper = float(metric["Slightly Undervalued"].split('-')[1].replace('%','')) / 100
                    fair_upper = float(metric["Fairly Valued"].split('-')[1].replace('%','')) / 100
                    sl_over_upper = float(metric["Slightly Overvalued"].split('-')[1].replace('%','')) / 100
                    formula = f"""=IF(ISBLANK(H{row}), "",
                        IF(H{row}<{sig_und}, "Significantly Undervalued",
                        IF(H{row}<{sl_und_upper}, "Slightly Undervalued",
                        IF(H{row}<={fair_upper}, "Fairly Valued",
                        IF(H{row}<{sl_over_upper}, "Slightly Overvalued",
                        "Significantly Overvalued")))))"""
                except Exception as e:
                    # If parsing fails, set a default assessment
                    formula = f"""=IF(ISBLANK(H{row}), "", "Assessment")"""
        else:
            # Handle numerical metrics
            if metric["Significantly Undervalued"].strip().upper() == "N/A":
                # Skip "Significantly Undervalued" and start with "Slightly Undervalued"
                try:
                    sl_und_upper = float(metric["Slightly Undervalued"].split('-')[1])
                    fair_upper = float(metric["Fairly Valued"].split('-')[1])
                    sl_over_upper = float(metric["Slightly Overvalued"].split('-')[1])
                    formula = f"""=IF(ISBLANK(H{row}), "",
                        IF(H{row}<{sl_und_upper}, "Slightly Undervalued",
                        IF(H{row}<{fair_upper}, "Fairly Valued",
                        IF(H{row}<{sl_over_upper}, "Slightly Overvalued",
                        "Significantly Overvalued"))))"""
                except Exception as e:
                    # If parsing fails, set a default assessment
                    formula = f"""=IF(ISBLANK(H{row}), "", "Assessment")"""
            else:
                # Regular formula with all categories
                try:
                    sig_und = float(metric["Significantly Undervalued"].replace('<','').replace('>',''))
                    sl_und_upper = float(metric["Slightly Undervalued"].split('-')[1])
                    fair_upper = float(metric["Fairly Valued"].split('-')[1])
                    sl_over_upper = float(metric["Slightly Overvalued"].split('-')[1])
                    formula = f"""=IF(ISBLANK(H{row}), "",
                        IF(H{row}<{sig_und}, "Significantly Undervalued",
                        IF(H{row}<{sl_und_upper}, "Slightly Undervalued",
                        IF(H{row}<{fair_upper}, "Fairly Valued",
                        IF(H{row}<{sl_over_upper}, "Slightly Overvalued",
                        "Significantly Overvalued")))))"""
                except Exception as e:
                    # If parsing fails, set a default assessment
                    formula = f"""=IF(ISBLANK(H{row}), "", "Assessment")"""
        
        # Assign the formula to the assessment cell
        assessment_cell.value = formula

# Step 2: Apply conditional formatting and adjust column widths
for industry, metrics in industry_data.items():
    try:
        ws = wb[industry]
    except KeyError:
        print(f"Error: Worksheet '{industry}' does not exist.")
        continue  # Skip to the next industry
    
    assessment_col = "I"
    for row in range(2, len(metrics) + 2):
        cell = f"{assessment_col}{row}"
        # Apply conditional formatting based on text
        for category, color in category_colors.items():
            if category == "N/A":
                continue  # Skip N/A for conditional formatting
            # Create a rule to apply the fill color if the cell text matches the category
            rule = FormulaRule(formula=[f'${assessment_col}${row}="{category}"'],
                               fill=PatternFill(start_color=color, end_color=color, fill_type="solid"))
            ws.conditional_formatting.add(cell, rule)
    
    # Adjust column widths for better readability
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20

# Save the workbook to an Excel file
file_path = "Stock_Valuation_Tool.xlsx"
wb.save(file_path)

print(f"Excel file '{file_path}' has been created successfully.")